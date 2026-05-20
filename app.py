import install_browsers  # installs Playwright Chromium on first run
import streamlit as st
import asyncio
import json
import re
import time
from datetime import datetime
from playwright.async_api import async_playwright
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

st.set_page_config(
    page_title="SIM Deal Auditor",
    page_icon="📱",
    layout="wide"
)

# ── Styling ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@700;800&display=swap');

html, body, [class*="css"] { font-family: 'DM Mono', monospace; }
h1, h2, h3 { font-family: 'Syne', sans-serif; }

.stApp { background: #0e0e0e; color: #e8e8e8; }

.block-container { padding: 2rem 2.5rem; max-width: 1400px; }

.hero {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    border: 1px solid #1e4d8c;
    border-radius: 12px;
    padding: 2.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.hero::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -20%;
    width: 400px;
    height: 400px;
    background: radial-gradient(circle, rgba(30,77,140,0.3) 0%, transparent 70%);
}
.hero h1 { font-size: 2.4rem; margin: 0 0 0.5rem; color: #fff; letter-spacing: -1px; }
.hero p { color: #8899bb; margin: 0; font-size: 0.85rem; }

.status-box {
    background: #111827;
    border: 1px solid #1f2937;
    border-radius: 8px;
    padding: 1rem 1.2rem;
    margin: 0.4rem 0;
    font-size: 0.8rem;
    color: #6b7280;
}
.status-box.active { border-color: #3b82f6; color: #93c5fd; }
.status-box.done { border-color: #10b981; color: #6ee7b7; }
.status-box.error { border-color: #ef4444; color: #fca5a5; }

.deal-count {
    font-family: 'Syne', sans-serif;
    font-size: 3rem;
    font-weight: 800;
    color: #3b82f6;
}

.site-badge {
    display: inline-block;
    padding: 0.2rem 0.7rem;
    border-radius: 4px;
    font-size: 0.75rem;
    font-weight: 500;
    margin: 0.2rem;
}
.badge-uswitch { background: #1e3a5f; color: #60a5fa; }
.badge-msm { background: #1e3a2e; color: #34d399; }
.badge-ctm { background: #3a1e1e; color: #f87171; }

stButton > button {
    background: #1d4ed8;
    color: white;
    border: none;
    border-radius: 8px;
    padding: 0.7rem 2rem;
    font-family: 'DM Mono', monospace;
    font-size: 0.85rem;
    cursor: pointer;
    transition: background 0.2s;
}
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
NETWORKS = ["Three", "Vodafone", "SMARTY", "TalkMobile", "VOXI", "iD Mobile", "Lebara", "GiffGaff", "Spusu", "O2", "Sky"]

SITES = {
    "uSwitch": {
        "url": "https://www.uswitch.com/mobiles/sim-only-deals/",
        "badge_class": "badge-uswitch"
    },
    "MoneySuperMarket": {
        "url": "https://www.moneysupermarket.com/mobile-phones/sim-only-deals/",
        "badge_class": "badge-msm"
    },
    "CompareTheMarket": {
        "url": "https://www.comparethemarket.com/mobile-phones/sim-only/",
        "badge_class": "badge-ctm"
    }
}

# Colour thresholds for GB (green = best, amber = mid, red = worst)
def gb_color(gb_val, all_vals_at_price):
    if not all_vals_at_price or gb_val is None:
        return None
    valid = [v for v in all_vals_at_price if v is not None]
    if not valid:
        return None
    mx = max(valid)
    mn = min(valid)
    if mx == mn:
        return "00B050"  # all same → green
    ratio = (gb_val - mn) / (mx - mn)
    if ratio >= 0.66:
        return "00B050"   # green
    elif ratio >= 0.33:
        return "FFC000"   # amber
    else:
        return "FF0000"   # red

# ── Scraping ──────────────────────────────────────────────────────────────────

async def scrape_uswitch(page, contract_months):
    """Scrape uSwitch SIM only deals"""
    deals = []
    try:
        await page.goto("https://www.uswitch.com/mobiles/sim-only-deals/", timeout=30000)
        await page.wait_for_timeout(3000)

        # Filter by contract length if possible
        try:
            if contract_months == 1:
                btn = page.locator("text=30 day").first
            elif contract_months == 12:
                btn = page.locator("text=12 month").first
            else:
                btn = page.locator("text=24 month").first
            await btn.click(timeout=3000)
            await page.wait_for_timeout(2000)
        except Exception:
            pass

        # Try to load more deals
        for _ in range(3):
            try:
                load_more = page.locator("text=Load more, text=Show more, button:has-text('more')").first
                await load_more.click(timeout=2000)
                await page.wait_for_timeout(1500)
            except Exception:
                break

        content = await page.content()
        deals = parse_deals_from_html(content, "uSwitch", contract_months)
    except Exception as e:
        st.session_state.scrape_log.append(f"uSwitch error: {e}")
    return deals


async def scrape_moneysupermarket(page, contract_months):
    deals = []
    try:
        await page.goto("https://www.moneysupermarket.com/mobile-phones/sim-only-deals/", timeout=30000)
        await page.wait_for_timeout(3000)

        # Try contract filter
        try:
            if contract_months == 1:
                sel = "30 days"
            elif contract_months == 12:
                sel = "12 months"
            else:
                sel = "24 months"
            btn = page.locator(f"text={sel}").first
            await btn.click(timeout=3000)
            await page.wait_for_timeout(2000)
        except Exception:
            pass

        for _ in range(3):
            try:
                load_more = page.locator("button:has-text('Show more'), button:has-text('Load more')").first
                await load_more.click(timeout=2000)
                await page.wait_for_timeout(1500)
            except Exception:
                break

        content = await page.content()
        deals = parse_deals_from_html(content, "MoneySuperMarket", contract_months)
    except Exception as e:
        st.session_state.scrape_log.append(f"MSM error: {e}")
    return deals


async def scrape_comparethemarket(page, contract_months):
    deals = []
    try:
        await page.goto("https://www.comparethemarket.com/mobile-phones/sim-only/", timeout=30000)
        await page.wait_for_timeout(4000)

        try:
            if contract_months == 1:
                sel = "1 month"
            elif contract_months == 12:
                sel = "12 months"
            else:
                sel = "24 months"
            btn = page.locator(f"text={sel}").first
            await btn.click(timeout=3000)
            await page.wait_for_timeout(2000)
        except Exception:
            pass

        for _ in range(3):
            try:
                load_more = page.locator("button:has-text('Show more'), button:has-text('Load more'), button:has-text('See more')").first
                await load_more.click(timeout=2000)
                await page.wait_for_timeout(1500)
            except Exception:
                break

        content = await page.content()
        deals = parse_deals_from_html(content, "CompareTheMarket", contract_months)
    except Exception as e:
        st.session_state.scrape_log.append(f"CTM error: {e}")
    return deals


def parse_deals_from_html(html, source, contract_months):
    """
    Parse deal data from raw HTML.
    Looks for patterns: price (£X), data (XGB / Unlimited), network name.
    Falls back to regex scanning of visible text.
    """
    deals = []
    # Remove scripts/styles for cleaner parsing
    clean = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r'<style[^>]*>.*?</style>', '', clean, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r'<[^>]+>', ' ', clean)
    clean = re.sub(r'\s+', ' ', clean)

    # Find all price mentions and nearby GB data
    # Pattern: £X.XX ... (N GB | Unlimited) ... (Network name)
    price_pattern = r'£(\d+(?:\.\d{2})?)'
    gb_pattern = r'(\d+)\s*GB|(\bUnlimited\b)'
    network_pattern = '|'.join(re.escape(n) for n in NETWORKS)

    # Sliding window approach: find prices, then look forward for GB
    price_matches = list(re.finditer(price_pattern, clean))

    seen = set()
    for pm in price_matches:
        price_val = float(pm.group(1))
        if price_val < 4 or price_val > 25:
            continue  # outside our range

        window = clean[pm.start(): pm.start() + 300]

        # Find GB
        gb_match = re.search(gb_pattern, window, re.IGNORECASE)
        if gb_match:
            if gb_match.group(2):  # Unlimited
                gb = "Unlimited"
                gb_num = 99999
            else:
                gb = int(gb_match.group(1))
                gb_num = gb
        else:
            continue  # no data found, skip

        # Find network
        net_match = re.search(network_pattern, window, re.IGNORECASE)
        network = net_match.group(0).strip() if net_match else "Unknown"

        # Deduplicate
        key = (source, network, price_val, str(gb))
        if key in seen:
            continue
        seen.add(key)

        deals.append({
            "source": source,
            "network": network,
            "price": price_val,
            "gb": gb,
            "gb_num": gb_num,
            "contract": contract_months,
        })

    return deals


async def run_scrape(contract_lengths, price_min, price_max):
    """Main async scrape orchestrator"""
    all_deals = []
    async with async_playwright() as p:
        import shutil, os, glob
        # Try multiple known Chromium locations across Render/Railway/Docker
        candidates = [
            "/usr/bin/chromium",
            "/usr/bin/chromium-browser",
            "/usr/bin/google-chrome",
            "/snap/bin/chromium",
        ]
        # Also check Playwright's own cache
        pw_cache = glob.glob("/root/.cache/ms-playwright/chromium-*/chrome-linux/chrome")
        candidates = pw_cache + candidates
        system_chromium = next((c for c in candidates if os.path.exists(c)), None)
        if not system_chromium:
            system_chromium = shutil.which("chromium") or shutil.which("chromium-browser")
        launch_kwargs = dict(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage", "--disable-gpu", "--single-process"]
        )
        if system_chromium:
            launch_kwargs["executable_path"] = system_chromium
        browser = await p.chromium.launch(**launch_kwargs)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900}
        )

        for months in contract_lengths:
            label = {1: "1M", 12: "12M", 24: "24M"}[months]
            st.session_state.scrape_log.append(f"▶ Scraping {label} contracts...")

            page = await context.new_page()

            # uSwitch
            st.session_state.current_site = f"uSwitch ({label})"
            deals = await scrape_uswitch(page, months)
            all_deals.extend(deals)
            st.session_state.scrape_log.append(f"  ✓ uSwitch: {len(deals)} deals found")

            # MoneySuperMarket
            st.session_state.current_site = f"MoneySuperMarket ({label})"
            deals = await scrape_moneysupermarket(page, months)
            all_deals.extend(deals)
            st.session_state.scrape_log.append(f"  ✓ MoneySuperMarket: {len(deals)} deals found")

            # CompareTheMarket
            st.session_state.current_site = f"CompareTheMarket ({label})"
            deals = await scrape_comparethemarket(page, months)
            all_deals.extend(deals)
            st.session_state.scrape_log.append(f"  ✓ CompareTheMarket: {len(deals)} deals found")

            await page.close()

        await browser.close()

    # Filter by price range
    all_deals = [d for d in all_deals if price_min <= d["price"] <= price_max]
    return all_deals


# ── Excel Export ──────────────────────────────────────────────────────────────

def build_excel(all_deals, contract_lengths):
    wb = Workbook()
    wb.remove(wb.active)

    SOURCES = ["uSwitch", "MoneySuperMarket", "CompareTheMarket"]
    CONTRACT_LABELS = {1: "1M", 12: "12M", 24: "24M"}
    PRICE_RANGE = range(5, 21)

    header_fills = {
        "uSwitch":           PatternFill("solid", fgColor="1F497D"),
        "MoneySuperMarket":  PatternFill("solid", fgColor="17375E"),
        "CompareTheMarket":  PatternFill("solid", fgColor="243F60"),
    }
    white_font = Font(color="FFFFFF", bold=True, size=9)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    green_fill  = PatternFill("solid", fgColor="00B050")
    amber_fill  = PatternFill("solid", fgColor="FFC000")
    red_fill    = PatternFill("solid", fgColor="FF0000")
    grey_fill   = PatternFill("solid", fgColor="F2F2F2")

    def get_fill(color_hex):
        return PatternFill("solid", fgColor=color_hex)

    for months in contract_lengths:
        label = CONTRACT_LABELS[months]
        ws = wb.create_sheet(title=f"{label} Market")

        # Build lookup: source → network → price → gb
        lookup = {}
        for d in all_deals:
            if d["contract"] != months:
                continue
            src = d["source"]
            net = d["network"]
            price = int(d["price"])
            gb = d["gb"]
            if src not in lookup:
                lookup[src] = {}
            if net not in lookup[src]:
                lookup[src][net] = {}
            lookup[src][net][price] = gb

        # Layout: 3 tables side by side with gap columns
        # Each table: Price col + N network cols
        networks_present = NETWORKS  # always show all columns

        table_width = 1 + len(networks_present)  # Price + networks
        gap = 1
        starts = [1, table_width + gap + 1, 2 * (table_width + gap) + 1]

        for t_idx, source in enumerate(SOURCES):
            col_start = starts[t_idx]

            # Source header (merged)
            ws.merge_cells(
                start_row=1, start_column=col_start,
                end_row=1, end_column=col_start + table_width - 1
            )
            cell = ws.cell(row=1, column=col_start, value=f"{source} {label} Market")
            cell.fill = header_fills[source]
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Column headers
            ws.cell(row=2, column=col_start, value="Price").font = Font(bold=True, size=8)
            ws.cell(row=2, column=col_start).fill = grey_fill

            for n_idx, network in enumerate(networks_present):
                c = col_start + 1 + n_idx
                cell = ws.cell(row=2, column=c, value=network)
                cell.font = Font(bold=True, size=8)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = grey_fill
                ws.column_dimensions[get_column_letter(c)].width = 10

            ws.column_dimensions[get_column_letter(col_start)].width = 7

            # Data rows
            for r_idx, price in enumerate(PRICE_RANGE):
                row = 3 + r_idx

                # Price cell
                price_cell = ws.cell(row=row, column=col_start, value=f"£{price}")
                price_cell.font = Font(bold=True, size=8)
                price_cell.alignment = Alignment(horizontal="center")

                # Collect all GB values at this price across networks for colour scaling
                gb_vals_at_price = []
                for network in networks_present:
                    gb = lookup.get(source, {}).get(network, {}).get(price)
                    if gb is not None and gb != "Unlimited":
                        try:
                            gb_vals_at_price.append(int(gb))
                        except Exception:
                            pass

                for n_idx, network in enumerate(networks_present):
                    c = col_start + 1 + n_idx
                    gb = lookup.get(source, {}).get(network, {}).get(price)
                    cell = ws.cell(row=row, column=c)
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=8)
                    cell.border = border

                    if gb is None:
                        cell.value = ""
                    elif gb == "Unlimited":
                        cell.value = "Unlimited"
                        cell.fill = green_fill
                        cell.font = Font(size=8, color="FFFFFF")
                    else:
                        cell.value = gb
                        # Colour based on relative value
                        color = gb_color(int(gb), gb_vals_at_price)
                        if color:
                            cell.fill = get_fill(color)
                            cell.font = Font(size=8, color="FFFFFF" if color in ("FF0000", "00B050") else "000000")

            ws.row_dimensions[1].height = 18
            ws.row_dimensions[2].height = 15

        ws.freeze_panes = "B3"

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum["A1"] = "SIM Deal Audit"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws_sum["A2"].font = Font(size=9, color="888888")
    ws_sum["A4"] = f"Total deals found: {len(all_deals)}"
    ws_sum["A5"] = f"Sources: {', '.join(SOURCES)}"
    ws_sum["A6"] = f"Contract lengths: {', '.join(CONTRACT_LABELS[m] for m in contract_lengths)}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── UI ────────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="hero">
    <h1>📱 SIM Deal Auditor</h1>
    <p>Automated price comparison across uSwitch · MoneySuperMarket · CompareTheMarket</p>
</div>
""", unsafe_allow_html=True)

# Session state init
if "scrape_log" not in st.session_state:
    st.session_state.scrape_log = []
if "deals" not in st.session_state:
    st.session_state.deals = []
if "current_site" not in st.session_state:
    st.session_state.current_site = ""
if "running" not in st.session_state:
    st.session_state.running = False

# ── Controls ──────────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns([2, 1, 1])

with col1:
    st.markdown("#### Contract Lengths")
    c1, c2, c3 = st.columns(3)
    inc_1m  = c1.checkbox("1 Month",   value=True)
    inc_12m = c2.checkbox("12 Months", value=True)
    inc_24m = c3.checkbox("24 Months", value=True)

with col2:
    st.markdown("#### Price Range")
    price_min = st.number_input("Min £", min_value=4, max_value=20, value=5)
    price_max = st.number_input("Max £", min_value=5, max_value=25, value=20)

with col3:
    st.markdown("#### Run Audit")
    st.write("")
    run_btn = st.button("🔍 Start Audit", use_container_width=True, type="primary")
    if st.session_state.deals:
        excel_data = build_excel(
            st.session_state.deals,
            [m for m, inc in [(1, inc_1m), (12, inc_12m), (24, inc_24m)] if inc]
        )
        st.download_button(
            "⬇ Download Excel",
            data=excel_data,
            file_name=f"sim_audit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.divider()

# ── Run scrape ────────────────────────────────────────────────────────────────
if run_btn:
    contract_lengths = [m for m, inc in [(1, inc_1m), (12, inc_12m), (24, inc_24m)] if inc]
    if not contract_lengths:
        st.warning("Please select at least one contract length.")
    else:
        st.session_state.scrape_log = []
        st.session_state.deals = []
        st.session_state.running = True

        progress_placeholder = st.empty()
        log_placeholder = st.empty()

        with progress_placeholder.container():
            st.markdown("### 🔄 Scraping in progress...")
            prog = st.progress(0)

        total_steps = len(contract_lengths) * 3
        step = 0

        async def run_all():
            return await run_scrape(contract_lengths, price_min, price_max)

        # Run async scraper
        deals = asyncio.run(run_all())
        st.session_state.deals = deals
        st.session_state.running = False

        progress_placeholder.empty()
        st.success(f"✅ Audit complete — {len(deals)} deals collected")

# ── Log & Results ─────────────────────────────────────────────────────────────
if st.session_state.scrape_log:
    with st.expander("📋 Scrape Log", expanded=False):
        for line in st.session_state.scrape_log:
            st.text(line)

if st.session_state.deals:
    st.markdown(f"### Results")
    df = pd.DataFrame(st.session_state.deals)

    # Summary stats
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Deals", len(df))
    m2.metric("uSwitch",          len(df[df.source == "uSwitch"]))
    m3.metric("MoneySuperMarket", len(df[df.source == "MoneySuperMarket"]))
    m4.metric("CompareTheMarket", len(df[df.source == "CompareTheMarket"]))

    st.markdown("#### Deal Table Preview")
    display_df = df[["source", "network", "price", "gb", "contract"]].copy()
    display_df.columns = ["Source", "Network", "Price (£)", "Data", "Contract (M)"]
    display_df = display_df.sort_values(["Source", "Price (£)", "Network"])
    st.dataframe(display_df, use_container_width=True, height=400)

    # Build and offer Excel download
    contract_lengths = [m for m, inc in [(1, inc_1m), (12, inc_12m), (24, inc_24m)] if inc]
    excel_data = build_excel(st.session_state.deals, contract_lengths)
    st.download_button(
        "⬇ Download Full Excel Report",
        data=excel_data,
        file_name=f"sim_audit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.markdown("""
    <div style="text-align:center; padding: 4rem 2rem; color: #374151;">
        <div style="font-size: 3rem; margin-bottom: 1rem;">📡</div>
        <div style="font-family: 'Syne', sans-serif; font-size: 1.2rem; color: #6b7280;">
            Select your options above and click Start Audit
        </div>
        <div style="font-size: 0.8rem; color: #4b5563; margin-top: 0.5rem;">
            Scraping uSwitch · MoneySuperMarket · CompareTheMarket simultaneously
        </div>
    </div>
    """, unsafe_allow_html=True)
